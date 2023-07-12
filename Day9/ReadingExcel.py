import openpyxl

wb= openpyxl.load_workbook("C:\\Users\\sds01798\\PycharmProjects\\ RupeshTraining\\PythonBasics\\Day9\\Book1.xlsx")
print("Active Sheet is:",wb.active.title)
#create object of any sheet on which you want to work
sh=wb['Sheet1']
print(sh.title)
print(sh['A4'].value)
print(sh['B2'].value)
c1=sh.cell(1,1)
print(c1.value)
c2=sh.cell(column=1,row=3)
print(c2.value)
print(c2.row)
print(c2.column)

